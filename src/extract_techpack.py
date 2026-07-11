"""TechPack BOM PDF 를 분석해 고객 Excel 양식(12컬럼)으로 추출한다.

처리 흐름
1. 설정 로드 및 클라이언트 생성
2. TechPack BOM 커스텀 분석기(techpack_bom) 보장(없으면 생성)
3. PDF 를 ``begin_analyze_binary`` 로 분석 → raw JSON 저장
4. 구조화 필드 정제(JSON) 저장
5. (소재 × 컬러웨이) 언피벗 → 고객 양식 Excel 생성

사용법:
    python -m src.extract_techpack "TechPack.pdf"
    python -m src.extract_techpack "TechPack.pdf" --out output/techpack --recreate-analyzer
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

from azure.ai.contentunderstanding import ContentUnderstandingClient
from azure.ai.contentunderstanding.models import ContentAnalyzer, ContentAnalyzerConfig

from .config import build_credential, load_settings
from .create_analyzer import (
    _get_analyzer,
    _model_mismatches,
    _schema_name_mismatch,
    _wait_for_analyzer_ready,
)
from .extract_work_order import (
    _analyze_binary_raw,
    _has_extracted_value,
    _simplify_fields,
    _unwrap_result,
    _validate_input_file,
    _validate_output_paths,
    _write_json_file,
)
from .retry import call_with_propagation_retry
from .techpack_schema import SCHEMA_NAME as TECHPACK_SCHEMA_NAME
from .techpack_schema import build_techpack_schema

# 고객 결과 Excel 의 12개 컬럼(순서 고정)
EXCEL_COLUMNS = [
    "Style",
    "COLORWAY",
    "Section",
    "WEB# / ID#",
    "DESCRIPTION",
    "QUALITY DETAILS",
    "SUPPLIER",
    "ARTICLE#",
    "ITEM COLOR",
    "UOM",
    "Item Price",
    "COMPONENT",
]

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_INVALID_EXCEL_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\uD800-\uDFFF\uFFFE\uFFFF]")
_PRICE_PREFIX = re.compile(
    r"^\s*(?:(?P<code>[A-Za-z]{3})\s+)?"
    r"(?P<symbol>[$€£¥₩₹₽])?\s*"
    r"(?P<amount>[+-]?[\d.,]+)\s+"
    r"(?P<unit>\S+)"
)


def _text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _unique_colorways(value: Any) -> list[str]:
    """문서 컬러웨이를 문자열 목록으로 정규화하고 중복을 제거한다."""

    if not isinstance(value, list):
        return []
    result: list[str] = []
    seen: set[str] = set()
    for item in value:
        name = _text(item)
        normalized = _norm_colorway(name)
        if name and normalized not in seen:
            seen.add(normalized)
            result.append(name)
    return result


def _excel_safe(value: Any) -> Any:
    """추출 문자열이 Excel 수식으로 실행되지 않도록 텍스트로 고정한다."""

    if isinstance(value, str):
        value = _INVALID_EXCEL_CHARACTERS.sub("", value)
        if value.startswith(_FORMULA_PREFIXES):
            return f"'{value}"
    return value


def ensure_techpack_analyzer(
    client: ContentUnderstandingClient, settings, *, recreate: bool = False
) -> str:
    """TechPack BOM 분석기가 존재하도록 보장하고 analyzer_id 를 반환한다."""

    analyzer_id = settings.techpack_analyzer_id
    existing = _get_analyzer(client, analyzer_id)

    if existing is not None and not recreate:
        existing = _wait_for_analyzer_ready(
            client,
            analyzer_id,
            existing,
            recreate_hint="TechPack 명령에 --recreate-analyzer 를 추가해 교체하세요.",
        )
        schema_mismatch = _schema_name_mismatch(existing, TECHPACK_SCHEMA_NAME)
        if schema_mismatch:
            raise SystemExit(
                f"기존 분석기 '{analyzer_id}' 의 스키마가 TechPack 스키마와 다릅니다: {schema_mismatch}\n"
                "TECHPACK_ANALYZER_ID 를 고유하게 지정하거나 TechPack 명령에 "
                "--recreate-analyzer 를 추가해 교체하세요."
            )
        mismatches = _model_mismatches(existing, settings)
        if mismatches:
            raise SystemExit(
                f"기존 분석기 '{analyzer_id}' 의 모델 설정이 현재 .env 와 다릅니다: "
                f"{', '.join(mismatches)}\n"
                "스키마 변경 내용을 확인한 뒤 TechPack 명령에 --recreate-analyzer 를 한 번 추가하세요."
            )
        print(f"분석기 '{analyzer_id}' 가 이미 존재합니다. (재사용)")
        return analyzer_id
    if existing is not None and recreate:
        print(f"기존 분석기 '{analyzer_id}' 를 새 스키마/모델로 교체합니다...")

    print(f"커스텀 분석기 '{analyzer_id}' 를 생성합니다...")
    analyzer = ContentAnalyzer(
        base_analyzer_id="prebuilt-document",
        description="의류 TechPack BILL OF MATERIALS 데이터 추출 분석기",
        config=ContentAnalyzerConfig(
            enable_ocr=True,
            enable_layout=True,
            enable_formula=False,
            estimate_field_source_and_confidence=True,
            return_details=True,
        ),
        field_schema=build_techpack_schema(),
        models={
            "completion": settings.completion_model,
            "embedding": settings.embedding_model,
        },
    )
    call_with_propagation_retry(
        lambda: client.begin_create_analyzer(
            analyzer_id=analyzer_id,
            resource=analyzer,
            allow_replace=True,
        ).result(),
        action="모델 배포 또는 역할",
    )
    print(f"분석기 '{analyzer_id}' 생성 완료.")
    return analyzer_id


def analyze_pdf(
    pdf_path: Path,
    *,
    recreate_analyzer: bool,
    raw_output_path: Path | None = None,
) -> dict[str, Any]:
    """PDF 를 분석하고 raw/structured 결과 dict 를 반환한다."""

    _validate_input_file(pdf_path)
    if raw_output_path is not None:
        _validate_output_paths(pdf_path, [raw_output_path])
    settings = load_settings()
    client = ContentUnderstandingClient(endpoint=settings.endpoint, credential=build_credential(settings))
    analyzer_id = ensure_techpack_analyzer(client, settings, recreate=recreate_analyzer)

    print(f"'{pdf_path.name}' 분석 중... (페이지 수에 따라 시간이 걸릴 수 있음)")
    raw_json = _analyze_binary_raw(client, analyzer_id=analyzer_id, input_path=pdf_path)
    if raw_output_path is not None:
        _write_json_file(raw_output_path, raw_json)
    raw_hint = f" 원본 응답은 {raw_output_path}에 저장했습니다." if raw_output_path is not None else ""

    result_body = _unwrap_result(raw_json)
    contents = result_body.get("contents", []) or []
    if not contents:
        raise RuntimeError(f"분석 결과에 contents 가 없습니다. 원본 응답과 분석기 설정을 확인하세요.{raw_hint}")
    fields: dict[str, Any] = {}
    markdown = None
    for content in contents:
        fields = _simplify_fields(content.get("fields", {}) or {}, with_confidence=False)
        markdown = content.get("markdown")
        break

    if not _has_extracted_value(fields):
        raise RuntimeError(
            f"분석은 완료됐지만 구조화 fields 가 비어 있습니다. TechPack 스키마와 모델 매핑을 확인하세요.{raw_hint}"
        )

    structured = {
        "sourceFile": pdf_path.name,
        "analyzerId": result_body.get("analyzerId", analyzer_id),
        "fields": fields,
        "markdown": markdown,
    }
    return {"raw": raw_json, "structured": structured}


def _norm_colorway(name: str) -> str:
    """컬러웨이 매칭용 정규화(공백/대소문자 무시)."""
    return " ".join(name.split()).strip().upper()


def _clean_price(price: str) -> str:
    """단가에서 통화+숫자+단위만 남기고 뒤의 SP24/LIST 등 태그 제거."""
    price = _text(price)
    match = _PRICE_PREFIX.match(price)
    if match:
        code = f"{match.group('code')} " if match.group("code") else ""
        symbol = match.group("symbol") or ""
        return f"{code}{symbol}{match.group('amount')} {match.group('unit')}"
    return price


def _clean_description(desc: str, article_number: str = "") -> str:
    """ARTICLE# 과 명확히 일치하는 선두 코드만 분리하고 나머지는 보존한다."""

    desc = _text(desc)
    article_number = _text(article_number)
    if article_number and desc[: len(article_number)].casefold() == article_number.casefold():
        boundary = desc[len(article_number) : len(article_number) + 1]
        if not boundary or boundary.isspace() or boundary == ":":
            return desc[: len(article_number)]
    return desc


def _resolve_components(materials: list[dict[str, Any]]) -> list[str]:
    """명시적으로 추출된 COMPONENT를 신뢰하고 빈 값만 앞 행에서 이어받는다."""

    resolved: list[str] = []
    last = ""
    for m in materials:
        comp = _text(m.get("component")) if isinstance(m, dict) else ""
        if comp:
            last = comp
        resolved.append(last)
    return resolved


def _has_material_data(material: dict[str, Any]) -> bool:
    """그룹 헤더/빈 객체가 아닌 실제 BOM 자재 데이터인지 확인한다."""

    detail_keys = (
        "section",
        "webId",
        "description",
        "qualityDetails",
        "supplier",
        "articleNumber",
        "price",
        "uom",
    )
    if any(_text(material.get(key)) for key in detail_keys):
        return True
    variants = material.get("colorVariants") or []
    return isinstance(variants, list) and any(
        isinstance(variant, dict) and _text(variant.get("itemColor")) for variant in variants
    )


def unpivot_rows(fields: dict[str, Any]) -> list[dict[str, str]]:
    """추출 fields 를 (소재 × 컬러웨이) 로 언피벗해 Excel 행 dict 리스트로 반환."""

    style = _text(fields.get("style"))
    doc_colorways = _unique_colorways(fields.get("colorways"))
    materials = fields.get("materials") or []
    if not isinstance(materials, list):
        return []
    resolved_components = _resolve_components(materials)
    material_rows = [
        (material, resolved_components[index])
        for index, material in enumerate(materials)
        if isinstance(material, dict) and _has_material_data(material)
    ]
    materials = [material for material, _ in material_rows]
    components = [component for _, component in material_rows]

    colorway_order = list(doc_colorways)
    colorway_seen = {_norm_colorway(name) for name in colorway_order}
    for mat in materials:
        if not isinstance(mat, dict):
            continue
        variants = mat.get("colorVariants") or []
        if not isinstance(variants, list):
            continue
        for variant in variants:
            if not isinstance(variant, dict):
                continue
            colorway = _text(variant.get("colorway"))
            normalized = _norm_colorway(colorway)
            if colorway and normalized not in colorway_seen:
                colorway_seen.add(normalized)
                colorway_order.append(colorway)

    rows: list[dict[str, str]] = []
    for idx, mat in enumerate(materials):
        if not isinstance(mat, dict):
            continue
        variants = mat.get("colorVariants") or []
        if not isinstance(variants, list):
            variants = []
        # 컬러웨이 -> itemColor 매핑
        by_cw: dict[str, str] = {}
        for v in variants:
            if not isinstance(v, dict):
                continue
            cw = _text(v.get("colorway"))
            if not cw:
                continue
            normalized = _norm_colorway(cw)
            by_cw[normalized] = _text(v.get("itemColor"))

        for cw in colorway_order:
            item_color = by_cw.get(_norm_colorway(cw), "")
            article_number = _text(mat.get("articleNumber"))
            rows.append(
                {
                    "Style": style,
                    "COLORWAY": cw,
                    "Section": _text(mat.get("section")),
                    "WEB# / ID#": _text(mat.get("webId")),
                    "DESCRIPTION": _clean_description(_text(mat.get("description")), article_number),
                    "QUALITY DETAILS": _text(mat.get("qualityDetails")),
                    "SUPPLIER": _text(mat.get("supplier")),
                    "ARTICLE#": article_number,
                    "ITEM COLOR": item_color,
                    "UOM": _text(mat.get("uom")),
                    "Item Price": _clean_price(_text(mat.get("price"))),
                    "COMPONENT": components[idx],
                }
            )
    return rows


def write_excel(rows: list[dict[str, str]], out_xlsx: Path) -> None:
    """언피벗 행을 고객 양식(12컬럼) Excel 로 저장한다."""

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "추출 결과"

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    top_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(EXCEL_COLUMNS)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = border

    for row in rows:
        ws.append([_excel_safe(row.get(col, "")) for col in EXCEL_COLUMNS])

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(EXCEL_COLUMNS)):
        for c in r:
            c.alignment = top_align
            c.border = border

    widths = [11, 22, 18, 11, 18, 34, 22, 18, 18, 7, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def _output_path(base: Path, extension: str) -> Path:
    """점이 포함된 베이스 파일명도 보존해 출력 확장자를 붙인다."""

    return base.parent / f"{base.name}{extension}"


def main() -> None:
    parser = argparse.ArgumentParser(description="TechPack BOM PDF -> 고객 양식 Excel 추출")
    parser.add_argument("pdf", type=Path, help="TechPack PDF 경로")
    parser.add_argument("--out", type=Path, default=None, help="출력 베이스 경로(확장자 제외). 기본 output/<파일명>")
    parser.add_argument("--recreate-analyzer", action="store_true", help="기존 분석기를 새 스키마/모델로 교체")
    args = parser.parse_args()

    _validate_input_file(args.pdf)

    base = args.out or (Path("output") / args.pdf.stem)
    raw_path = _output_path(base, ".raw.json")
    json_path = _output_path(base, ".json")
    xlsx_path = _output_path(base, ".xlsx")
    try:
        _validate_output_paths(args.pdf, [raw_path, json_path, xlsx_path])
    except ValueError as exc:
        parser.error(str(exc))
    base.parent.mkdir(parents=True, exist_ok=True)

    result = analyze_pdf(
        args.pdf,
        recreate_analyzer=args.recreate_analyzer,
        raw_output_path=raw_path,
    )

    _write_json_file(json_path, result["structured"])

    fields = result["structured"].get("fields") or {}
    materials_value = fields.get("materials") or []
    materials = materials_value if isinstance(materials_value, list) else []
    rows = unpivot_rows(fields)
    if not rows:
        sys.exit("TechPack 소재/컬러웨이 행을 만들 수 없습니다. 정제 JSON의 materials와 colorVariants를 확인하세요.")
    write_excel(rows, xlsx_path)
    colorways = list(dict.fromkeys(row["COLORWAY"] for row in rows))

    print(f"\n원본 JSON   : {raw_path}")
    print(f"정제 JSON   : {json_path}")
    print(f"결과 Excel  : {xlsx_path}")
    print(f"\nStyle       : {fields.get('style')!r}")
    print(f"컬러웨이({len(colorways)}): {colorways}")
    print(f"소재 행      : {len(materials)}건")
    print(f"언피벗 행    : {len(rows)}건 (= 소재 × 컬러웨이)")


if __name__ == "__main__":
    main()
