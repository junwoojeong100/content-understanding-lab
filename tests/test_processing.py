from __future__ import annotations

import io
import json
import os
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from azure.core.exceptions import HttpResponseError
from openpyxl import load_workbook

from src.config import load_settings
from src.extract_techpack import (
    EXCEL_COLUMNS,
    _clean_description,
    _clean_price,
    _output_path,
    analyze_pdf as analyze_techpack,
    unpivot_rows,
    write_excel,
)
from src.extract_work_order import (
    _has_extracted_value,
    _simplify_field,
    _unwrap_result,
    _validate_output_paths,
    analyze_pdf as analyze_work_order,
)
from src.create_analyzer import _model_mismatches, _schema_name_mismatch, _wait_for_analyzer_ready
from src.retry import call_with_propagation_retry


class ConfigTests(unittest.TestCase):
    def test_load_settings_rejects_disallowed_analyzer_id_characters(self) -> None:
        for analyzer_id in ("invalid/id", "trade-work-order"):
            env = {
                "CONTENTUNDERSTANDING_ENDPOINT": "https://example.services.ai.azure.com/",
                "WORK_ORDER_ANALYZER_ID": analyzer_id,
                "TECHPACK_ANALYZER_ID": "techpack_bom",
            }
            with self.subTest(analyzer_id=analyzer_id):
                with patch.dict(os.environ, env, clear=True), self.assertRaises(SystemExit):
                    load_settings()

    def test_load_settings_accepts_periods_and_underscores_in_analyzer_ids(self) -> None:
        env = {
            "CONTENTUNDERSTANDING_ENDPOINT": "https://example.services.ai.azure.com/",
            "WORK_ORDER_ANALYZER_ID": "trade.work_order",
            "TECHPACK_ANALYZER_ID": "techpack_bom",
        }
        with patch.dict(os.environ, env, clear=True):
            settings = load_settings()
        self.assertEqual(settings.analyzer_id, "trade.work_order")
        self.assertEqual(settings.techpack_analyzer_id, "techpack_bom")

    def test_load_settings_uses_current_model_defaults(self) -> None:
        env = {
            "CONTENTUNDERSTANDING_ENDPOINT": "https://example.services.ai.azure.com/",
            "WORK_ORDER_ANALYZER_ID": "trade_work_order",
            "TECHPACK_ANALYZER_ID": "techpack_bom",
        }
        with patch.dict(os.environ, env, clear=True):
            settings = load_settings()
        self.assertEqual(settings.completion_model, "gpt-5.2")
        self.assertEqual(settings.embedding_model, "text-embedding-3-large")

    def test_load_settings_rejects_duplicate_analyzer_ids(self) -> None:
        env = {
            "CONTENTUNDERSTANDING_ENDPOINT": "https://example.services.ai.azure.com/",
            "WORK_ORDER_ANALYZER_ID": "shared",
            "TECHPACK_ANALYZER_ID": "shared",
        }
        with patch.dict(os.environ, env, clear=True), self.assertRaises(SystemExit):
            load_settings()


class WorkOrderProcessingTests(unittest.TestCase):
    def test_simplify_nested_fields_with_confidence(self) -> None:
        field = {
            "type": "array",
            "valueArray": [
                {
                    "type": "object",
                    "valueObject": {
                        "itemCode": {
                            "type": "string",
                            "valueString": "A-100",
                            "confidence": 0.91,
                        }
                    },
                }
            ],
        }
        self.assertEqual(
            _simplify_field(field, with_confidence=True),
            [{"itemCode": {"value": "A-100", "confidence": 0.91}}],
        )

    def test_unwraps_lro_and_direct_results(self) -> None:
        result = {"contents": []}
        self.assertIs(_unwrap_result(result), result)
        self.assertEqual(_unwrap_result({"result": result}), result)

    def test_rejects_output_that_overwrites_input(self) -> None:
        with self.assertRaises(ValueError):
            _validate_output_paths(Path("source.xlsx"), [Path("source.xlsx")])

    def test_rejects_hard_link_output_that_overwrites_input(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.pdf"
            alias = Path(tmp) / "alias.pdf"
            source.write_bytes(b"%PDF-test")
            os.link(source, alias)
            with self.assertRaises(ValueError):
                _validate_output_paths(source, [alias])

    def test_rejects_output_artifacts_that_alias_each_other(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.pdf"
            first = Path(tmp) / "first.json"
            second = Path(tmp) / "second.json"
            source.write_bytes(b"%PDF-test")
            first.write_text("{}", encoding="utf-8")
            os.link(first, second)
            with self.assertRaises(ValueError):
                _validate_output_paths(source, [first, second])

    def test_saves_raw_response_before_structured_validation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.pdf"
            raw_path = Path(tmp) / "result.raw.json"
            input_path.write_bytes(b"%PDF-test")
            settings = SimpleNamespace(endpoint="https://example.services.ai.azure.com/", analyzer_id="trade_work_order")
            raw_response = {"contents": []}
            with (
                patch("src.extract_work_order.load_settings", return_value=settings),
                patch("src.extract_work_order.build_credential", return_value=object()),
                patch("src.extract_work_order.ContentUnderstandingClient"),
                patch("src.extract_work_order._analyze_binary_raw", return_value=raw_response),
                self.assertRaisesRegex(RuntimeError, "result.raw.json"),
            ):
                analyze_work_order(
                    input_path,
                    with_confidence=False,
                    ensure=False,
                    recreate_analyzer=False,
                    raw_output_path=raw_path,
                )
            self.assertEqual(json.loads(raw_path.read_text(encoding="utf-8")), raw_response)

    def test_programmatic_work_order_rejects_raw_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.pdf"
            input_path.write_bytes(b"%PDF-test")
            with self.assertRaises(ValueError):
                analyze_work_order(
                    input_path,
                    with_confidence=False,
                    ensure=False,
                    recreate_analyzer=False,
                    raw_output_path=input_path,
                )

    def test_detects_existing_analyzer_model_mismatch(self) -> None:
        analyzer = SimpleNamespace(models={"completion": "gpt-4.1", "embedding": "text-embedding-3-large"})
        settings = SimpleNamespace(completion_model="gpt-5.2", embedding_model="text-embedding-3-large")
        self.assertEqual(
            _model_mismatches(analyzer, settings),
            ["completion='gpt-4.1' (기대 'gpt-5.2')"],
        )

    def test_detects_existing_analyzer_schema_mismatch(self) -> None:
        analyzer = SimpleNamespace(field_schema=SimpleNamespace(name="trade_work_order_schema"))
        self.assertEqual(
            _schema_name_mismatch(analyzer, "techpack_bom_schema"),
            "field_schema.name='trade_work_order_schema' (기대 'techpack_bom_schema')",
        )

    def test_rejects_non_ready_existing_analyzer(self) -> None:
        analyzer = SimpleNamespace(status="failed")
        with self.assertRaisesRegex(SystemExit, "failed"):
            _wait_for_analyzer_ready(
                SimpleNamespace(),
                "trade_work_order",
                analyzer,
                recreate_hint="recreate",
                delay_seconds=0,
            )

    def test_rejects_multiple_contents_when_all_fields_are_empty(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.pdf"
            raw_path = Path(tmp) / "result.raw.json"
            input_path.write_bytes(b"%PDF-test")
            settings = SimpleNamespace(endpoint="https://example.services.ai.azure.com/", analyzer_id="trade_work_order")
            raw_response = {
                "contents": [
                    {"fields": {"a": {"type": "string"}}},
                    {"fields": {"b": {"type": "array", "valueArray": []}}},
                ]
            }
            with (
                patch("src.extract_work_order.load_settings", return_value=settings),
                patch("src.extract_work_order.build_credential", return_value=object()),
                patch("src.extract_work_order.ContentUnderstandingClient"),
                patch("src.extract_work_order._analyze_binary_raw", return_value=raw_response),
                self.assertRaisesRegex(RuntimeError, "구조화 fields"),
            ):
                analyze_work_order(
                    input_path,
                    with_confidence=False,
                    ensure=False,
                    recreate_analyzer=False,
                    raw_output_path=raw_path,
                )

    def test_meaningful_value_detection_preserves_zero_and_false(self) -> None:
        self.assertFalse(_has_extracted_value({"a": None, "b": [], "confidence": 0.9}))
        self.assertTrue(_has_extracted_value({"a": 0}))
        self.assertTrue(_has_extracted_value({"a": False}))
        self.assertFalse(_has_extracted_value({"value": None, "confidence": 0.9}))


class RetryTests(unittest.TestCase):
    def test_retries_only_known_propagation_error(self) -> None:
        attempts = 0

        def operation() -> str:
            nonlocal attempts
            attempts += 1
            if attempts == 1:
                raise HttpResponseError(message="DeploymentIdNotFound")
            return "ok"

        with redirect_stdout(io.StringIO()):
            result = call_with_propagation_retry(
                operation,
                action="test",
                max_attempts=2,
                delay_seconds=0,
            )
        self.assertEqual(result, "ok")
        self.assertEqual(attempts, 2)

    def test_does_not_retry_unrelated_http_error(self) -> None:
        attempts = 0

        def operation() -> None:
            nonlocal attempts
            attempts += 1
            raise HttpResponseError(message="BadRequest")

        with self.assertRaises(HttpResponseError):
            call_with_propagation_retry(
                operation,
                action="test",
                max_attempts=3,
                delay_seconds=0,
            )
        self.assertEqual(attempts, 1)


class TechPackProcessingTests(unittest.TestCase):
    def test_cleaners_preserve_expected_values(self) -> None:
        self.assertEqual(_clean_description("FVF4895QD: RECYCLED STRETCH", "FVF4895QD"), "FVF4895QD")
        self.assertEqual(_clean_description("DT5613 MSPORT 150", "DT5613"), "DT5613")
        self.assertEqual(_clean_description("FVF4895QD: RECYCLED STRETCH"), "FVF4895QD: RECYCLED STRETCH")
        self.assertEqual(_clean_description("No code material name"), "No code material name")
        self.assertEqual(
            _clean_description("PACKING TRIM inclusive of: CARTON LABEL"),
            "PACKING TRIM inclusive of: CARTON LABEL",
        )
        self.assertEqual(_clean_description("210D RECYCLED NYLON"), "210D RECYCLED NYLON")
        self.assertEqual(_clean_description("1000 DENIER CORDURA"), "1000 DENIER CORDURA")
        self.assertEqual(
            _clean_description("3M: SCOTCHLITE REFLECTIVE TRANSFER"),
            "3M: SCOTCHLITE REFLECTIVE TRANSFER",
        )
        self.assertEqual(_clean_description("300GSM FLEECE FABRIC"), "300GSM FLEECE FABRIC")
        self.assertEqual(_clean_description("4-WAY STRETCH FABRIC"), "4-WAY STRETCH FABRIC")
        self.assertEqual(_clean_description("210D RECYCLED NYLON", "210D"), "210D")
        self.assertEqual(_clean_price("$3.200 yd SP24"), "$3.200 yd")
        self.assertEqual(_clean_price("€3.200 m LIST"), "€3.200 m")
        self.assertEqual(_clean_price("USD 3.200 yd SP24"), "USD 3.200 yd")
        self.assertEqual(_clean_price("£ 10.50 ea LIST"), "£10.50 ea")

    def test_output_path_preserves_dotted_base_name(self) -> None:
        base = Path("output/TechPack.v2")
        self.assertEqual(_output_path(base, ".xlsx"), Path("output/TechPack.v2.xlsx"))
        self.assertEqual(_output_path(base, ".raw.json"), Path("output/TechPack.v2.raw.json"))

    def test_scalar_colorways_fall_back_to_variant_order(self) -> None:
        rows = unpivot_rows(
            {
                "style": "S1",
                "colorways": "BLACK 001",
                "materials": [
                    {
                        "component": "Fabric",
                        "section": "Woven",
                        "colorVariants": [{"colorway": "BLACK 001", "itemColor": "BLACK"}],
                    }
                ],
            }
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["COLORWAY"], "BLACK 001")

    def test_partial_document_colorways_include_all_material_variants(self) -> None:
        rows = unpivot_rows(
            {
                "style": "S1",
                "colorways": ["BLACK 001"],
                "materials": [
                    {
                        "component": "Fabric",
                        "section": "Woven",
                        "colorVariants": [
                            {"colorway": "BLACK 001", "itemColor": "BLACK"},
                            {"colorway": "NAVY 092", "itemColor": "NAVY"},
                        ],
                    }
                ],
            }
        )
        self.assertEqual([row["COLORWAY"] for row in rows], ["BLACK 001", "NAVY 092"])

    def test_saves_techpack_raw_response_before_structured_validation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.pdf"
            raw_path = Path(tmp) / "result.raw.json"
            input_path.write_bytes(b"%PDF-test")
            settings = SimpleNamespace(
                endpoint="https://example.services.ai.azure.com/",
                techpack_analyzer_id="techpack_bom",
            )
            raw_response = {"contents": []}
            with (
                patch("src.extract_techpack.load_settings", return_value=settings),
                patch("src.extract_techpack.build_credential", return_value=object()),
                patch("src.extract_techpack.ContentUnderstandingClient"),
                patch("src.extract_techpack.ensure_techpack_analyzer", return_value="techpack_bom"),
                patch("src.extract_techpack._analyze_binary_raw", return_value=raw_response),
                self.assertRaisesRegex(RuntimeError, "result.raw.json"),
            ):
                analyze_techpack(
                    input_path,
                    recreate_analyzer=False,
                    raw_output_path=raw_path,
                )
            self.assertEqual(json.loads(raw_path.read_text(encoding="utf-8")), raw_response)

    def test_programmatic_techpack_rejects_raw_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.pdf"
            input_path.write_bytes(b"%PDF-test")
            with self.assertRaises(ValueError):
                analyze_techpack(
                    input_path,
                    recreate_analyzer=False,
                    raw_output_path=input_path,
                )

    def test_unpivot_rows_and_excel_columns(self) -> None:
        fields = {
            "style": "SW002394",
            "colorways": ["BLACK 001", "DEEP NAVY 092"],
            "materials": [
                {
                    "component": "Fabric",
                    "section": "Woven / Plain",
                    "webId": "LSVO",
                    "description": "FVF4895QD: fabric",
                    "qualityDetails": "86% Polyester",
                    "supplier": "Everest Textile",
                    "articleNumber": "FVF4895QD",
                    "price": "$3.200 yd SP24",
                    "uom": "yd",
                    "colorVariants": [
                        {"colorway": "BLACK 001", "itemColor": "BLACK - 001"},
                        {"colorway": "DEEP NAVY 092", "itemColor": "DEEP NAVY - 092"},
                    ],
                },
                {
                    "component": "",
                    "section": "Woven / Plain",
                    "webId": "LT6I",
                    "description": "DT5613 MSPORT 150",
                    "qualityDetails": "100% Polyester",
                    "supplier": "Designer Textiles",
                    "articleNumber": "DT5613",
                    "price": "$10.480 m LIST",
                    "uom": "m",
                    "colorVariants": [
                        {"colorway": "BLACK 001", "itemColor": "(Not Colorable)"},
                        {"colorway": "DEEP NAVY 092", "itemColor": "(Not Colorable)"},
                    ],
                },
            ],
        }
        rows = unpivot_rows(fields)
        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[0]["COMPONENT"], "Fabric")
        self.assertEqual(rows[2]["COMPONENT"], "Fabric")
        self.assertEqual(rows[0]["DESCRIPTION"], "FVF4895QD")
        self.assertEqual(rows[2]["DESCRIPTION"], "DT5613")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "result.xlsx"
            write_excel(rows, path)
            workbook = load_workbook(path, read_only=True)
            worksheet = workbook["추출 결과"]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(headers, EXCEL_COLUMNS)
            self.assertEqual(worksheet.max_row, 5)
            workbook.close()

    def test_component_equal_to_section_is_preserved(self) -> None:
        rows = unpivot_rows(
            {
                "style": "S1",
                "colorways": ["BLACK"],
                "materials": [
                    {
                        "component": "Fabric",
                        "section": "Woven",
                        "colorVariants": [{"colorway": "BLACK", "itemColor": "BLACK"}],
                    },
                    {
                        "component": "Thread",
                        "section": "Thread",
                        "colorVariants": [{"colorway": "BLACK", "itemColor": "BLACK"}],
                    },
                ],
            }
        )
        self.assertEqual([row["COMPONENT"] for row in rows], ["Fabric", "Thread"])

    def test_empty_material_objects_do_not_create_blank_rows(self) -> None:
        rows = unpivot_rows(
            {
                "style": "S1",
                "colorways": ["BLACK"],
                "materials": [
                    {
                        "component": "Fabric",
                        "description": None,
                        "colorVariants": [],
                    }
                ],
            }
        )
        self.assertEqual(rows, [])

    def test_component_only_header_is_carried_to_next_material(self) -> None:
        rows = unpivot_rows(
            {
                "style": "S1",
                "colorways": ["BLACK"],
                "materials": [
                    {"component": "Fabric"},
                    {
                        "component": "",
                        "description": "RECYCLED NYLON",
                        "colorVariants": [{"colorway": "BLACK", "itemColor": "BLACK"}],
                    },
                ],
            }
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["COMPONENT"], "Fabric")

    def test_excel_values_are_not_written_as_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "safe.xlsx"
            write_excel([{"Style": "=1+1", "COLORWAY": "@SUM(A1:A2)"}], path)
            workbook = load_workbook(path, data_only=False, read_only=True)
            worksheet = workbook["추출 결과"]
            self.assertEqual(worksheet["A2"].data_type, "s")
            self.assertEqual(worksheet["A2"].value, "'=1+1")
            self.assertEqual(worksheet["B2"].data_type, "s")
            workbook.close()

    def test_excel_values_remove_invalid_xml_characters(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "safe.xlsx"
            write_excel([{"Style": "bad\x00text", "COLORWAY": "\x00=1+1"}], path)
            workbook = load_workbook(path, data_only=False, read_only=True)
            worksheet = workbook["추출 결과"]
            self.assertEqual(worksheet["A2"].value, "badtext")
            self.assertEqual(worksheet["B2"].value, "'=1+1")
            self.assertEqual(worksheet["B2"].data_type, "s")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
